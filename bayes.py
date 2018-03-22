# -*- coding:utf-8 -*-
import tushare as ts
from openpyxl import load_workbook
import datetime
import csv

class TrainVo(object):
	def __init__(self):
		self.data = None


class BayesStock:
	def __init__(self):
		self.startDay = datetime.datetime.strptime('2017/12/1', '%Y/%m/%d')
		self.endDay = datetime.datetime.strptime('2017/12/2', '%Y/%m/%d')
		self.trainData = []

	#获取所有股票列表,只需要调用一次存入Excel
	def getTotalList(self):
		df = ts.get_stock_basics()
		df.to_excel('./data/total_list.xlsx',encoding = 'utf-8')

	#获取训练数据,取上一个月数据存入csv
	def getTrainData(self):
		lw = load_workbook('./data/total_list.xlsx')
		sheet = lw.active
		for row in sheet['A']:
			df = ts.get_k_data(code=row.value,start='2017-12-01',end='2017-12-31',ktype='D')
			df.to_csv('./data/'+row.value+'.csv',index=False)

	def getOneDayData(self):
		day = datetime.timedelta(days=1)
		curDay = self.startDay
		nextDay = self.startDay + day
		delta = self.endDay - nextDay
		if delta.days <= 0:
			print(delta.days)
		print(nextDay)
		lw = load_workbook('./data/total_list.xlsx')
		sheet = lw.active
		# for row in sheet['A']:
		# 	with open(row.value + '.csv') as csvfile:
		# 		reader = csv.DictReader(csvfile)

		nextIndex = nextDay.strftime('%Y-%m-%d')
		print('nextDay:'+index)
		with open('./data/300133' + '.csv') as csvfile:
			reader = csv.DictReader(csvfile)
			for row in reader:
				if row['date'] == nextIndex:
					vo = TrainVo()
					vo.data = row
					self.trainData.append(vo)


bayes = BayesStock()
bayes.getOneDayData()
# bayes.getTotalList()
# bayes.getTrainData()